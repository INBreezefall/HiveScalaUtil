#-*- coding:utf-8 -*-
import os
from openpyxl import load_workbook

filename="数据表.xlsx"
wb=load_workbook(filename)
sheets = wb.sheetnames
# print(type(sheets))
# print(sheets[0],sheets[1])

for i in range(len(sheets)):
    #获取某sheet页内容
    booksheet=wb[sheets[i]]
    #获取行/列
    cols = booksheet.columns
    # rows = booksheet.rows
    #取得数据库table名
    tableN=str(sheets[i])
    #页面总数据
    list1=[]
    #需要的4列数据
    sjx = []
    sjt = []
    lon = []
    bz = []
    #赋值总数据
    for col in cols:
        list1.append([row.value for row in col][1:])
        pass
    pass
    #赋值4列变量
    sjx=list1[1]
    sjx1=list1[1].copy()
    bz=list1[2]
    sjt=list1[3]
    lon=list1[4]

    #拼接语句
    tableE=tableN.split(",")[1]
    tableC=tableN.split(",")[0]
    top="CREATE TABLE "+tableE+"("
    body=""
    comment=""
    for i in range(len(sjt)):
        sjx1[i]=""
        if(sjx[i] in sjx1):
            continue
        if(sjt[i] == "C"):
            t="varchar2("
            if ("None"==str(lon[i])):
                t=t+"1000)"
            else:
                t=t+str(lon[i])+")"
        elif(sjt[i] == "N"):
            t="NUMBER("
            if("，" in str(lon[i])):
                f=str(lon[i]).split("，")[0]
                s=str(lon[i]).split("，")[1]
                t=t+f+","+s+")"
            elif("," in str(lon[i])):
                f=str(lon[i]).split(",")[0]
                s=str(lon[i]).split(",")[1]
                t=t+f+","+s+")"
            elif(str(lon[i])=="None"):
                t=t+"12)"
            else:
                t=t+str(lon[i])+",0)"
        elif(sjt[i] == "M"):
            t="varchar2("
            if ("None"==str(lon[i])):
                t=t+"10)"
            else:
                t=t+str(lon[i])+")"
        elif(sjt[i] == "B"):
            t="varchar2("
            if ("None"==str(lon[i])):
                t=t+"1000)"
            else:
                t=t+str(lon[i])+")"
        elif(sjt[i] == "T"):
            t="varchar2(2000)"
        else:
            exit("警告：%s.%s %s数据类型%s有误"%(tableE,sjt[i],i,sjt[i]))
        pass
        if("/" in str(sjx[i])):
            sjx[i]=str(sjx[i]).split("/")[0]
        body=body+sjx[i]+" "+t+" "+""+","
        comment=comment+"comment on column "+tableE+"."+sjx[i]+" is "+"'%s'"%(bz[i])+";"
        pass
    pass
    sql=top+body[:-1]+");"
    comment=comment+"comment on table "+tableE+" is '%s';"%tableC
    fileName="数据表SQL.txt"
    with open(fileName,"a",encoding="utf-8") as fp:
        #写入数据
        fp.write(sql+"\n")
        fp.write(comment+"\n\n")
        print("*"*20+"保存ORACLE_SQL语句完毕"+"*"*20)
